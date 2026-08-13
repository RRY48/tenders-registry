# -*- coding: utf-8 -*-
import openpyxl, json
SRC = "../тендеры/Реестр тендеров 2026 (новая структура).xlsx"
OUT = "data.json"
ws = openpyxl.load_workbook(SRC).active
hdr = [c.value for c in ws[1]]
out = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[3]:
        continue
    out.append({h: (str(r[i]) if r[i] is not None else "") for i, h in enumerate(hdr) if h})
json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("OK", len(out))